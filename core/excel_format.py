from pathlib import Path
from copy import copy
import pandas as pd
from openpyxl.utils import get_column_letter

def autosize_columns(worksheet, dataframe: pd.DataFrame, max_width: int = 42) -> None:
    """
    Set Excel column widths safely.

    Important:
    Do not use pandas Series.map(len) here because some pandas/Python versions
    can keep numeric objects as float and len(float) raises TypeError.
    """
    for idx, col in enumerate(dataframe.columns, start=1):
        max_len = len(str(col))

        for value in dataframe[col].tolist():
            if pd.isna(value):
                text = ""
            else:
                text = str(value)
            max_len = max(max_len, len(text))

        width = min(max_len + 2, max_width)
        worksheet.column_dimensions[worksheet.cell(row=1, column=idx).column_letter].width = width

def write_excel_openpyxl(
    register: pd.DataFrame,
    output_path: Path,
    register_sheet_name: str,
    summary_func,
) -> None:
    """
    Write output using openpyxl engine.
    This avoids the extra xlsxwriter dependency.
    """
    with pd.ExcelWriter(output_path, engine="openpyxl", date_format="yyyy-mm-dd", datetime_format="yyyy-mm-dd") as writer:
        register.to_excel(writer, sheet_name=register_sheet_name, index=False)
        summary_func(writer, register)

        workbook = writer.book

        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

            for cell in ws[1]:
                font = copy(cell.font)
                font.bold = True
                cell.font = font

                fill = copy(cell.fill)
                fill.fill_type = "solid"
                fill.fgColor = "D9EAF7"
                cell.fill = fill

            # Basic number formatting by header names.
            headers = {cell.value: cell.column for cell in ws[1]}
            for header, col_idx in headers.items():
                col_letter = ws.cell(row=1, column=col_idx).column_letter

                if header and "Date" in str(header) or header in ["Data", "NBP rate date", "Date of issue"]:
                    for row in range(2, ws.max_row + 1):
                        ws[f"{col_letter}{row}"].number_format = "yyyy-mm-dd"

                if header in [
                    "Netto PLN", "VAT PLN", "Brutto PLN",
                    "Net EUR source", "VAT EUR source", "Gross EUR source",
                    "Original net amount", "Original VAT amount", "Original gross amount",
                    "Netto_PLN", "VAT_PLN", "Brutto_PLN",
                ]:
                    for row in range(2, ws.max_row + 1):
                        ws[f"{col_letter}{row}"].number_format = '#,##0.00'

                if header == "NBP EUR/PLN rate":
                    for row in range(2, ws.max_row + 1):
                        ws[f"{col_letter}{row}"].number_format = '0.0000'

            # Autosize.
            if sheet_name == register_sheet_name:
                autosize_columns(ws, register)
            elif sheet_name == "Summary":
                summary_df = pd.read_excel(output_path, sheet_name="Summary") if False else None
                for col_idx in range(1, ws.max_column + 1):
                    ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 18
            elif sheet_name == "Problems":
                autosize_columns(ws, register, max_width=42)

def write_multi_sheet_excel_openpyxl(output_path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    """Write several DataFrames to one workbook with consistent simple formatting."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl", date_format="yyyy-mm-dd", datetime_format="yyyy-mm-dd") as writer:
        for sheet_name, df in sheets.items():
            safe_name = sheet_name[:31]
            df.to_excel(writer, sheet_name=safe_name, index=False)

        workbook = writer.book
        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

            for cell in ws[1]:
                font = copy(cell.font)
                font.bold = True
                cell.font = font
                fill = copy(cell.fill)
                fill.fill_type = "solid"
                fill.fgColor = "D9EAF7"
                cell.fill = fill

            headers = {cell.value: cell.column for cell in ws[1]}
            for header, col_idx in headers.items():
                col_letter = ws.cell(row=1, column=col_idx).column_letter
                header_text = str(header or "")
                if "Date" in header_text or header in ["Data", "NBP rate date", "Date of issue", "Period start", "Period end"]:
                    for row in range(2, ws.max_row + 1):
                        ws[f"{col_letter}{row}"].number_format = "yyyy-mm-dd"
                if header in [
                    "Netto PLN", "VAT PLN", "Brutto PLN",
                    "Net EUR source", "VAT EUR source", "Gross EUR source",
                    "Original net amount", "Original VAT amount", "Original gross amount",
                    "Netto_PLN", "VAT_PLN", "Brutto_PLN",
                ]:
                    for row in range(2, ws.max_row + 1):
                        ws[f"{col_letter}{row}"].number_format = "#,##0.00"
                if header == "NBP EUR/PLN rate":
                    for row in range(2, ws.max_row + 1):
                        ws[f"{col_letter}{row}"].number_format = "0.0000"

            for col_idx in range(1, ws.max_column + 1):
                max_len = 10
                for row_idx in range(1, min(ws.max_row, 200) + 1):
                    value = ws.cell(row=row_idx, column=col_idx).value
                    max_len = max(max_len, len(str(value)) if value is not None else 0)
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 46)
