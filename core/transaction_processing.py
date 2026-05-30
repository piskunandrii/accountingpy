from __future__ import annotations

import csv
import html
import re
import unicodedata
from datetime import datetime
from html.parser import HTMLParser
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ============================================================
# 1. CONFIG
# ============================================================

POSSIBLE_BASE_DIRS = [
    Path("/Users/andriiipiskun/Desktop/ISS-Chem sp. z o. o."),
    Path("/Users/andriiipiskun/Desktop"),
    Path("/Users/andriiipiskun/Library/Mobile Documents/com~apple~CloudDocs/Desktop/ISS-Chem sp. z o. o."),
    Path("/Users/andriiipiskun/Library/Mobile Documents/com~apple~CloudDocs/Desktop"),
]

TARGET_SHEETS = {
    ("MILLENNIUM", "EUR"): "EUR-MILL",
    ("MILLENNIUM", "PLN"): "PLN-MILL",
    ("ING", "EUR"): "EUR-ING",
    ("ING", "PLN"): "PLN-ING",
}

OUTPUT_HEADERS = [
    "Дата транзакції",
    "Контрагент",
    "Тип транзакції",
    "Номер рахунку контрагента",
    "Сума транзакції",
    "Сума після транзакції",
    "Оригінальний опис транзакції",
]

LEGAL_PHRASES = [
    "spolka z ograniczona odpowiedzialnoscia",
    "spółka z ograniczoną odpowiedzialnością",
    "limited liability company",
    "sp z o o",
    "sp zoo",
    "sp. z o.o.",
    "llc",
    "ltd",
    "s a",
    "sa",
    "b v",
    "bv",
    "inc",
]

ADDRESS_STOP_WORDS = {
    "ul",
    "aleja",
    "al",
    "gen",
    "antoniego",
    "chrusciela",
    "chruściela",
    "montera",
    "warszawa",
    "rotterdam",
    "delftseplein",
    "poltava",
    "banka",
    "ivana",
    "str",
    "apt",
    "room",
    "ua",
    "nl",
    "pl",
    "hoholya",
    "myrhorod",
    "myrhorodcity",
    "region",
    "city",
    "gumińska",
    "guminska",
    "usa",
    "http",
    "https",
}

DESCRIPTION_COUNTERPARTY_PATTERNS = [
    {
        "description_keys": ["sqsp", "squarespace", "squarespacec", "squaresp"],
        "account_name_keys": ["squarespace"],
        "fallback_name": "Squarespace INC",
    },
]


# ============================================================
# 2. BASIC NORMALIZATION HELPERS
# ============================================================

def strip_accents(text: Any) -> str:
    text = unicodedata.normalize("NFKD", str(text))
    return "".join(ch for ch in text if not unicodedata.combining(ch))


def clean_excel_text(x: Any) -> str:
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return ""

    s = str(x).strip()

    # Millennium exports some cells as Excel formulas like =T("text")
    m = re.match(r'^=T\("(.*)"\)$', s)
    if m:
        s = m.group(1)

    return s.strip()


def normalize_filename_for_search(name: Any) -> str:
    text = unicodedata.normalize("NFKD", str(name)).lower()
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^a-zа-яіїєґ0-9]+", "", text)
    return text


def compact_filename(text: Any) -> str:
    return re.sub(r"[^a-z0-9а-яіїєґ]+", "", strip_accents(str(text)).lower())


def loose_text_key(*texts: Any) -> str:
    s = strip_accents(" ".join("" if t is None else str(t) for t in texts).lower())
    return re.sub(r"[^a-z0-9]+", "", s)


def norm_account(x: Any) -> str:
    s = clean_excel_text(x)
    return re.sub(r"[^A-Za-z0-9]", "", s).upper()


def parse_amount(x: Any) -> Optional[float]:
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return None

    if isinstance(x, (int, float)):
        return float(x)

    s = str(x).strip().replace("\xa0", "").replace(" ", "")
    if not s or s.lower() == "nan":
        return None

    # Polish format: 1.234,56
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(",", ".")

    try:
        return float(s)
    except ValueError:
        return None


def parse_date(x: Any) -> Optional[datetime]:
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return None

    if isinstance(x, datetime):
        return x.replace(tzinfo=None)

    s = clean_excel_text(x)
    if not s:
        return None

    if re.match(r"^\d{1,2}/\d{1,2}/\d{4}$", s):
        dt = pd.to_datetime(s, errors="coerce", dayfirst=True)
    else:
        dt = pd.to_datetime(s, errors="coerce", dayfirst=False)

    if not pd.isna(dt):
        return dt.to_pydatetime().replace(tzinfo=None)

    return None


def normalize_cols(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out.columns = [
        re.sub(r"\s+", " ", str(c).replace("\xa0", " ")).strip()
        for c in out.columns
    ]
    return out


# ============================================================
# 3. FILE DISCOVERY
# ============================================================

def find_existing_base_dir() -> Path:
    for folder in POSSIBLE_BASE_DIRS:
        if not folder.exists() or not folder.is_dir():
            continue

        files = [
            p for p in folder.iterdir()
            if p.is_file() and p.suffix.lower() in [".xlsx", ".xls", ".csv"]
        ]

        if files:
            return folder

    raise FileNotFoundError(
        "Не знайшов папку з Excel/CSV файлами. "
        "Перевір, чи файли лежать на Desktop або iCloud Desktop."
    )


def find_master_file(folder: Path) -> Path:
    candidates: List[Path] = []

    for file_path in folder.iterdir():
        if not file_path.is_file():
            continue

        if file_path.name.startswith("~$"):
            continue

        if file_path.suffix.lower() != ".xlsx":
            continue

        normalized = normalize_filename_for_search(file_path.name)

        is_master = (
            ("історія" in normalized and "транзакц" in normalized)
            or ("istoriia" in normalized and "tranzak" in normalized)
            or ("istoria" in normalized and "tranzak" in normalized)
        )

        is_output_copy = (
            "zapowneno" in normalized
            or "заповнено" in normalized
            or "filled" in normalized
            or "updated" in normalized
        )

        if is_master and not is_output_copy:
            candidates.append(file_path)

    if not candidates:
        raise FileNotFoundError(
            f"Не знайдено головний Excel-файл у папці: {folder}. "
            "Назва має бути на кшталт 'Історія Транзакцій.xlsx'."
        )

    if len(candidates) > 1:
        print("Знайдено кілька можливих головних файлів:")
        for i, path in enumerate(candidates, start=1):
            print(f"{i}. {path}")

        raise RuntimeError(
            "Є кілька файлів, схожих на головний файл. "
            "Залиши в папці тільки один головний файл або уточни назву вручну."
        )

    return candidates[0]


def find_transaction_files(folder: Path, master_file: Path) -> List[Path]:
    result: List[Path] = []
    allowed_ext = {".xlsx", ".xls", ".csv"}

    print()
    print(f"Папка пошуку: {folder}")
    print("Excel/CSV файли, які бачить Python:")

    all_files: List[Path] = []

    for p in folder.iterdir():
        if p.is_file() and p.suffix.lower() in allowed_ext:
            all_files.append(p)
            print("-", repr(p.name), "=>", compact_filename(p.name))

    print()

    for p in all_files:
        if p.name.startswith("~$"):
            continue

        try:
            if p.resolve() == master_file.resolve():
                continue
        except FileNotFoundError:
            pass

        name_key = compact_filename(p.name)
        master_key = compact_filename(master_file.name)

        if name_key == master_key:
            continue

        # Skip previously generated copies and unrelated accounting workbooks.
        if (
            "zapowneno" in name_key
            or "заповнено" in p.name.lower()
            or "accountingtable" in name_key
        ):
            continue

        is_bank_file = (
            "historia" in name_key
            and ("operacji" in name_key or "transakcji" in name_key)
        )

        if is_bank_file:
            result.append(p)

    return sorted(result)


# ============================================================
# 4. READING BANK FILES
# ============================================================

class SimpleHTMLTableParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.tables: List[List[List[str]]] = []
        self.current_table: Optional[List[List[str]]] = None
        self.current_row: Optional[List[str]] = None
        self.current_cell: Optional[List[str]] = None
        self.in_cell = False

    def handle_starttag(self, tag: str, attrs: List[Tuple[str, Optional[str]]]) -> None:
        tag = tag.lower()
        if tag == "table":
            self.current_table = []
        elif tag == "tr" and self.current_table is not None:
            self.current_row = []
        elif tag in {"td", "th"} and self.current_row is not None:
            self.current_cell = []
            self.in_cell = True

    def handle_data(self, data: str) -> None:
        if self.in_cell and self.current_cell is not None:
            self.current_cell.append(data)

    def handle_endtag(self, tag: str) -> None:
        tag = tag.lower()
        if tag in {"td", "th"} and self.current_cell is not None and self.current_row is not None:
            text = html.unescape("".join(self.current_cell)).strip()
            text = re.sub(r"\s+", " ", text)
            self.current_row.append(text)
            self.current_cell = None
            self.in_cell = False
        elif tag == "tr" and self.current_row is not None and self.current_table is not None:
            if any(cell.strip() for cell in self.current_row):
                self.current_table.append(self.current_row)
            self.current_row = None
        elif tag == "table" and self.current_table is not None:
            if self.current_table:
                self.tables.append(self.current_table)
            self.current_table = None


def dataframe_from_html_fallback(path: Path) -> Optional[pd.DataFrame]:
    raw = path.read_bytes()

    for enc in ["utf-8", "cp1250", "windows-1250", "iso-8859-2", "latin1"]:
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            text = ""
    else:
        return None

    if "<table" not in text.lower():
        return None

    parser = SimpleHTMLTableParser()
    parser.feed(text)

    if not parser.tables:
        return None

    table = max(parser.tables, key=len)
    if len(table) < 2:
        return None

    headers = table[0]
    rows = table[1:]

    max_len = max(len(headers), *(len(r) for r in rows))
    headers = headers + [f"Column_{i}" for i in range(len(headers) + 1, max_len + 1)]

    normalized_rows = []
    for r in rows:
        normalized_rows.append(r + [""] * (max_len - len(r)))

    return normalize_cols(pd.DataFrame(normalized_rows, columns=headers[:max_len]))


def read_csv_flexible(path: Path) -> pd.DataFrame:
    raw = path.read_bytes()

    for enc in ["utf-8-sig", "utf-8", "cp1250", "windows-1250", "iso-8859-2", "latin1"]:
        try:
            sample = raw[:4096].decode(enc)
            encoding = enc
            break
        except UnicodeDecodeError:
            continue
    else:
        encoding = "latin1"
        sample = raw[:4096].decode(encoding, errors="ignore")

    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,|\t")
        sep = dialect.delimiter
    except Exception:
        sep = ";"

    return normalize_cols(pd.read_csv(path, sep=sep, dtype=str, encoding=encoding))


def read_bank_file(path: Path) -> pd.DataFrame:
    suffix = path.suffix.lower()

    if suffix == ".csv":
        return read_csv_flexible(path)

    # Many Millennium .xls files are actually HTML tables.
    try:
        tables = pd.read_html(path, encoding="utf-8")
        if tables:
            return normalize_cols(tables[0])
    except Exception:
        pass

    fallback_html = dataframe_from_html_fallback(path)
    if fallback_html is not None:
        return fallback_html

    # Try real Excel engines.
    if suffix == ".xlsx":
        return normalize_cols(pd.read_excel(path, dtype=str, engine="openpyxl"))

    try:
        return normalize_cols(pd.read_excel(path, dtype=str, engine="xlrd"))
    except Exception as e:
        raise ValueError(
            f"Не можу прочитати файл {path.name}. "
            "Файл має розширення .xls, але не читається ні як HTML, ні як Excel. "
            "Спробуй експортувати його з банку як CSV або XLSX."
        ) from e


def bank_from_df(df: pd.DataFrame) -> str:
    cols = {str(c).lower() for c in df.columns}

    if any("numer rachunku/karty" in c for c in cols):
        return "MILLENNIUM"

    if any("rachunek ing" in c for c in cols):
        return "ING"

    raise ValueError(f"Не можу визначити банк. Колонки: {list(df.columns)}")


# ============================================================
# 5. ACCOUNTS
# ============================================================

def name_tokens(text: Any) -> List[str]:
    if text is None or (isinstance(text, float) and pd.isna(text)):
        return []

    s = strip_accents(str(text).lower())
    s = re.sub(r"[^a-z0-9]+", " ", s)

    for phrase in LEGAL_PHRASES:
        s = s.replace(strip_accents(phrase.lower()), " ")

    tokens: List[str] = []

    for token in s.split():
        if token in ADDRESS_STOP_WORDS:
            continue
        if token.isdigit() or len(token) <= 1:
            continue
        if sum(ch.isdigit() for ch in token) >= 2:
            continue
        tokens.append(token)

    return tokens


def normalize_name_key(text: Any) -> str:
    return "".join(name_tokens(text)[:5])


def load_accounts(master: Path) -> Tuple[Dict[str, Dict[str, Any]], List[Dict[str, Any]]]:
    wb = openpyxl.load_workbook(master, data_only=True)

    if "Accounts" not in wb.sheetnames:
        raise RuntimeError("У головному Excel немає аркуша 'Accounts'.")

    ws = wb["Accounts"]

    headers = [cell.value for cell in ws[1]]
    idx = {h: i for i, h in enumerate(headers) if h}

    required = ["Назва", "Банк", "Валюта", "Рахунок"]
    missing = [h for h in required if h not in idx]

    if missing:
        raise RuntimeError(f"На аркуші Accounts немає колонок: {missing}")

    accounts_by_num: Dict[str, Dict[str, Any]] = {}
    accounts: List[Dict[str, Any]] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[idx["Назва"]]
        bank = row[idx["Банк"]]
        curr = row[idx["Валюта"]]
        acct = row[idx["Рахунок"]]

        if not acct or not name:
            continue

        rec = {
            "name": str(name).strip(),
            "bank": str(bank).strip() if bank else "",
            "currency": str(curr).strip().upper() if curr else "",
            "account_raw": str(acct).strip(),
            "account_norm": norm_account(acct),
            "tokens": name_tokens(name),
            "name_key": normalize_name_key(name),
        }

        accounts.append(rec)

        if rec["account_norm"]:
            accounts_by_num[rec["account_norm"]] = rec

            # Millennium may export Polish account without PL prefix.
            if rec["account_norm"].startswith("PL"):
                accounts_by_num[rec["account_norm"][2:]] = rec

    return accounts_by_num, accounts


# ============================================================
# 6. COUNTERPARTY MATCHING
# ============================================================

def account_from_description(description: Any, accounts: List[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    desc_key = loose_text_key(description)

    if not desc_key:
        return None

    for pattern in DESCRIPTION_COUNTERPARTY_PATTERNS:
        if not any(k in desc_key for k in pattern["description_keys"]):
            continue

        for rec in accounts:
            rec_key = loose_text_key(rec.get("name", ""))
            if all(k in rec_key for k in pattern["account_name_keys"]):
                return rec

    return None


def fallback_name_from_description(description: Any) -> str:
    desc_key = loose_text_key(description)

    for pattern in DESCRIPTION_COUNTERPARTY_PATTERNS:
        if any(k in desc_key for k in pattern["description_keys"]):
            return str(pattern.get("fallback_name") or "")

    return ""


def identify_counterparty(
    name: Any,
    account: Any,
    accounts_by_num: Dict[str, Dict[str, Any]],
    accounts: List[Dict[str, Any]],
    description: Any = "",
) -> Tuple[str, Optional[Dict[str, Any]]]:

    # First, match by bank-provided counterparty name.
    tokens = set(name_tokens(name))

    if tokens:
        best = None
        best_score = 0
        normalized_input_name = normalize_name_key(name)

        for rec in accounts:
            rtokens = set(rec["tokens"])
            if not rtokens:
                continue

            common = tokens & rtokens

            if rtokens.issubset(tokens):
                score = 100 + len(rtokens)
            elif rec["name_key"] and (
                rec["name_key"] in normalized_input_name
                or normalized_input_name in rec["name_key"]
            ):
                score = 80 + len(common)
            else:
                score = len(common)

            if score > best_score and score >= 2:
                best = rec
                best_score = score

        if best:
            return best["name"], best

    # Second, match by counterparty bank account.
    acc = norm_account(account)

    if acc in accounts_by_num:
        rec = accounts_by_num[acc]
        return rec["name"], rec

    # Third, match known merchants by transaction description.
    desc_rec = account_from_description(description, accounts)
    if desc_rec:
        return desc_rec["name"], desc_rec

    fallback = fallback_name_from_description(description)
    if fallback:
        return fallback, None

    raw = clean_excel_text(name)
    return re.sub(r"\s+", " ", raw), None


# ============================================================
# 7. TRANSACTION TYPE CLASSIFICATION
# ============================================================

def is_known_merchant_description(description: Any) -> bool:
    return bool(fallback_name_from_description(description))


def is_commission_type(tx_type: Any) -> bool:
    return str(tx_type).startswith("комісія")


def bank_counterparty_name(account_rec: Dict[str, Any], bank: str) -> str:
    raw = (account_rec or {}).get("bank", "") or bank
    low = str(raw).lower()

    if "millennium" in low:
        return "Millennium Bank"

    if "ing" in low:
        return "ING Bank"

    return str(raw).strip() or bank


def bank_label_from_account(account_rec: Dict[str, Any], fallback_bank: str) -> str:
    b = str((account_rec or {}).get("bank", "")).lower()

    if "millennium" in b:
        return "MILLENNIUM"

    if "ing" in b:
        return "ING"

    return fallback_bank


def classify_transaction(counterparty: Any = "", description: Any = "", *texts: Any) -> str:
    all_texts = [description, *texts]
    s = strip_accents(" ".join("" if t is None else str(t) for t in all_texts).lower())
    compact = loose_text_key(*all_texts)
    cp_compact = loose_text_key(counterparty)

    # Known merchants such as Squarespace must not be treated as card commissions.
    if is_known_merchant_description(description):
        return "переказ"

    # Loan: only when description contains "umowa pożyczki" and counterparty is Andrii Piskun.
    if "umowapozyczki" in compact and (
        "andriipiskun" in cp_compact or "piskunandrii" in cp_compact
    ):
        return "позика"

    # Salary / remuneration, including typo variants.
    salary_markers = [
        "wynagrodz",
        "wynagodz",
        "wynagdz",
        "wynagdzenie",
        "wynagodzenie",
        "wynagrdodzenie",
        "wynagrdodzen",
        "salary",
        "listaplac",
        "placa",
        "umowazlecenia",
    ]

    if any(k in compact for k in salary_markers) or any(
        k in s for k in ["wynagrodz", "salary", "lista plac", "placa", "umowa zlecenia"]
    ):
        return "зарплата"

    # Millennium SCI fee descriptions.
    if "sci" in compact and any(
        k in compact for k in ["przekazsepa", "przelewzagraniczny"]
    ):
        return "комісія за переказ"

    if "sms" in s:
        return "комісія за SMS"

    words = set(re.findall(r"[a-z]+", s))

    if "pit" in words:
        return "податок PIT"

    if "cit" in words:
        return "податок CIT"

    if "zus" in words:
        return "податок ZUS"

    if any(k in s for k in ["wymiana waluty", "exchange", "currency"]):
        return "обмін валюти"

    if any(k in s for k in ["prowadzenie rachunku", "oplata za konto", "opłata za konto", "maintenance"]):
        return "комісія за ведення рахунку"

    if any(k in s for k in ["prowizja", "oplata", "opłata", "commission", "fee"]):
        if any(k in s for k in ["przelew", "transfer", "zagraniczny", "sepa"]):
            return "комісія за переказ"
        return "комісія за ведення рахунку"

    # Card fee only if it looks like an actual bank fee, not a card purchase merchant.
    if any(k in s for k in ["opłata za kartę", "oplata za karte", "card fee"]):
        return "комісія за карту"

    return "переказ"


# ============================================================
# 8. NORMALIZATION
# ============================================================

def normalize_file(
    path: Path,
    accounts_by_num: Dict[str, Dict[str, Any]],
    accounts: List[Dict[str, Any]],
) -> Dict[str, List[List[Any]]]:

    df = read_bank_file(path)
    bank = bank_from_df(df)

    rows_by_sheet: Dict[str, List[List[Any]]] = {
        sheet: [] for sheet in TARGET_SHEETS.values()
    }

    for _, r in df.iterrows():
        if bank == "ING":
            own_account = norm_account(r.get('rachunek ING "IBAN" (pl+26 znaków)'))
            date = parse_date(r.get("data transakcji"))
            curr = str(r.get("waluta operacji", "")).strip().upper()

            cp_acct_raw = r.get("rachunek kontrahenta IBAN (IBAN PL+ 26.znaków)", "")
            cp_acct = norm_account(cp_acct_raw)
            cp_name_raw = r.get("nazwa i adres kontrahenta", "")

            desc = clean_excel_text(r.get("nazwa transakcji i opis", ""))

            amount = parse_amount(r.get("kwota"))
            balance = parse_amount(r.get("saldo po transakcji"))

            tx_texts = [
                r.get("kod transakcji 4 znaki", ""),
                cp_name_raw,
            ]

        else:
            own_account = norm_account(r.get("Numer rachunku/karty"))
            date = parse_date(r.get("Data transakcji"))
            curr = str(r.get("Waluta", "")).strip().upper()

            cp_acct_raw = r.get("Na rachunek/ Z rachunku", "")
            cp_acct = norm_account(cp_acct_raw)
            cp_name_raw = r.get("Odbiorca/Nadawca", "")

            desc = clean_excel_text(r.get("Opis", ""))

            debit = parse_amount(r.get("Obciążenie"))
            credit = parse_amount(r.get("Uznanie"))
            amount = credit if credit is not None else debit

            balance = parse_amount(r.get("Saldo"))

            tx_texts = [
                r.get("Rodzaj transakcji", ""),
                cp_name_raw,
            ]

        account_rec = accounts_by_num.get(own_account, {})
        bank_final = bank_label_from_account(account_rec, bank)
        curr_final = curr or account_rec.get("currency", "")

        sheet = TARGET_SHEETS.get((bank_final, curr_final))
        if not sheet:
            continue

        counterparty, cp_rec = identify_counterparty(
            cp_name_raw,
            cp_acct,
            accounts_by_num,
            accounts,
            desc,
        )

        cp_out = cp_acct

        if cp_rec and (
            not cp_out
            or cp_out == own_account
            or (own_account.startswith("PL") and cp_out == own_account[2:])
            or cp_out == cp_rec["account_norm"][2:]
        ):
            cp_out = cp_rec["account_norm"]

        # If known merchant from description has an account in Accounts, use it.
        if cp_rec and is_known_merchant_description(desc):
            cp_out = cp_rec["account_norm"]

        tx_type = classify_transaction(counterparty, desc, *tx_texts)

        # Bank commissions: counterparty is bank and account is own account.
        # But known merchants such as Squarespace are never treated as bank commissions.
        if is_commission_type(tx_type) and not is_known_merchant_description(desc):
            counterparty = bank_counterparty_name(account_rec, bank_final)
            cp_out = own_account

        rows_by_sheet[sheet].append([
            date,
            counterparty,
            tx_type,
            cp_out,
            amount,
            balance,
            desc,
        ])

    return rows_by_sheet


# ============================================================
# 9. MERGE INTO EXISTING MASTER FILE
# ============================================================

def normalize_cell_for_key(value: Any) -> str:
    if value is None:
        return ""

    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")

    if isinstance(value, (int, float)):
        return f"{float(value):.2f}"

    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)

    # Excel sometimes returns datetime as text after previous saves.
    dt = parse_date(text)
    if dt:
        return dt.strftime("%Y-%m-%d")

    amount = parse_amount(text)
    if amount is not None and re.search(r"\d", text):
        return f"{amount:.2f}"

    return text


def transaction_key(row: List[Any]) -> str:
    """
    Unique transaction key.

    row structure:
    0 Date
    1 Counterparty
    2 Transaction type
    3 Counterparty account
    4 Amount
    5 Balance after transaction
    6 Original bank description

    We intentionally do NOT include counterparty and transaction type in the key,
    because those are exactly the fields that may be improved by newer code rules.
    """

    date_key = normalize_cell_for_key(row[0])
    amount_key = normalize_cell_for_key(row[4])
    balance_key = normalize_cell_for_key(row[5])
    account_key = norm_account(row[3])
    description_key = loose_text_key(row[6])

    return "|".join([
        date_key,
        amount_key,
        balance_key,
        account_key,
        description_key,
    ])


def worksheet_row_to_list(ws: openpyxl.worksheet.worksheet.Worksheet, row_idx: int) -> List[Any]:
    values: List[Any] = []

    for col_idx in range(1, len(OUTPUT_HEADERS) + 1):
        values.append(ws.cell(row=row_idx, column=col_idx).value)

    return values


def is_empty_transaction_row(row: List[Any]) -> bool:
    return all(value is None or str(value).strip() == "" for value in row)


def normalize_existing_row_shape(row: List[Any]) -> List[Any]:
    """
    Ensures old rows fit the current 7-column structure.
    Useful if an older version of the file had fewer columns.
    """

    row = list(row)

    if len(row) < len(OUTPUT_HEADERS):
        row.extend([""] * (len(OUTPUT_HEADERS) - len(row)))

    if len(row) > len(OUTPUT_HEADERS):
        row = row[:len(OUTPUT_HEADERS)]

    return row


def merge_existing_and_new_rows(
    existing_rows: List[List[Any]],
    new_rows: List[List[Any]],
) -> List[List[Any]]:
    """
    Old rows are preserved.
    New rows are added.
    Existing transactions are replaced by the newly normalized version.

    This allows you to import bank files that start later, for example from 2026-05-01,
    without deleting transactions before that date from the master workbook.
    """

    merged_by_key: Dict[str, List[Any]] = {}

    for row in existing_rows:
        row = normalize_existing_row_shape(row)

        if is_empty_transaction_row(row):
            continue

        key = transaction_key(row)

        if key:
            merged_by_key[key] = row

    for row in new_rows:
        row = normalize_existing_row_shape(row)

        if is_empty_transaction_row(row):
            continue

        key = transaction_key(row)

        if key:
            # New normalized version wins.
            merged_by_key[key] = row

    return list(merged_by_key.values())


def sort_date_value(value: Any) -> datetime:
    if isinstance(value, datetime):
        return value

    parsed = parse_date(value)
    if parsed:
        return parsed

    # Rows without dates go last when sorting descending.
    return datetime.min


def apply_sheet_formatting(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(
        min_row=1,
        max_row=ws.max_row,
        max_col=len(OUTPUT_HEADERS),
    ):
        for cell in row:
            cell.border = Border(
                left=thin,
                right=thin,
                top=thin,
                bottom=thin,
            )
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    if ws.max_row > 1:
        for cell in ws["A"][1:]:
            cell.number_format = "yyyy-mm-dd"

        for cell in list(ws["E"][1:]) + list(ws["F"][1:]):
            cell.number_format = "#,##0.00"

    widths = [16, 34, 24, 34, 18, 20, 60]

    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{max(ws.max_row, 1)}"


def write_output(
    master: Path,
    output: Path,
    rows_by_sheet: Dict[str, List[List[Any]]],
) -> None:
    wb = openpyxl.load_workbook(master)

    for sheet_name in TARGET_SHEETS.values():
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)

        existing_rows: List[List[Any]] = []

        # Read old rows before clearing sheet.
        if ws.max_row >= 2:
            for row_idx in range(2, ws.max_row + 1):
                row_values = worksheet_row_to_list(ws, row_idx)
                row_values = normalize_existing_row_shape(row_values)

                if not is_empty_transaction_row(row_values):
                    existing_rows.append(row_values)

        new_rows = rows_by_sheet.get(sheet_name, [])

        merged_rows = merge_existing_and_new_rows(existing_rows, new_rows)

        # Sort from newest to oldest.
        merged_rows = sorted(
            merged_rows,
            key=lambda row: sort_date_value(row[0]),
            reverse=True,
        )

        if ws.max_row:
            ws.delete_rows(1, ws.max_row)

        ws.append(OUTPUT_HEADERS)

        for row in merged_rows:
            ws.append(row)

        apply_sheet_formatting(ws)

    wb.save(output)


# ============================================================
# 10. MAIN
# ============================================================

def main() -> None:
    base_dir = find_existing_base_dir()
    master_file = find_master_file(base_dir)
    output_file = master_file

    print(f"BASE_DIR: {base_dir}")
    print(f"MASTER_FILE: {master_file}")
    print(f"OUTPUT_FILE: {output_file}")

    accounts_by_num, accounts = load_accounts(master_file)
    bank_files = find_transaction_files(base_dir, master_file)

    print("Знайдені банківські файли:")
    for p in bank_files:
        print("-", p.name)

    if not bank_files:
        raise RuntimeError(
            "Не знайдено жодного банківського файлу. "
            "Подивись вище в секцію 'Excel/CSV файли, які бачить Python'."
        )

    aggregate: Dict[str, List[List[Any]]] = {
        sheet: [] for sheet in TARGET_SHEETS.values()
    }

    for path in bank_files:
        print()
        print(f"Обробляю файл: {path.name}")

        part = normalize_file(path, accounts_by_num, accounts)

        for sheet, rows in part.items():
            aggregate[sheet].extend(rows)

        print("Рядків з файлу:")
        for sheet, rows in part.items():
            print(f"  {sheet}: {len(rows)}")

    print()
    print("Кількість нових/оновлених нормалізованих рядків з банківських файлів:")
    for sheet, rows in aggregate.items():
        print(f"{sheet}: {len(rows)}")

    write_output(master_file, output_file, aggregate)

    print()
    print(f"Збережено в оригінальний файл: {output_file}")
    print("Режим запису: merge/update. Старі транзакції, яких немає в нових банківських файлах, не видаляються.")


if __name__ == "__main__":
    main()
