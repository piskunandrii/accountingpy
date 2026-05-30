import contextlib
import io
import threading
from datetime import datetime
from typing import Any
from tkinter import BOTH, LEFT, Tk, Toplevel, messagebox
from tkinter import ttk

import openpyxl

from app.registers_gui import RejestrApp
from core import transaction_processing

BALANCE_SHEETS = [
    ("MILLENNIUM", "EUR", "EUR-MILL"),
    ("MILLENNIUM", "PLN", "PLN-MILL"),
    ("ING", "EUR", "EUR-ING"),
    ("ING", "PLN", "PLN-ING"),
]


def run_app() -> None:
    root = Tk()
    MainApp(root)
    root.mainloop()


class MainApp:
    def __init__(self, root: Tk):
        self.root = root
        self.root.title("ISS-Chem бухгалтерський helper")
        self.root.geometry("640x330")
        self.root.resizable(False, False)
        self.registers_window: Toplevel | None = None

        self._build_ui()
        self.refresh_balances()

    def _build_ui(self) -> None:
        main = ttk.Frame(self.root, padding=18)
        main.pack(fill=BOTH, expand=True)

        actions = ttk.LabelFrame(main, text="Головне меню", padding=12)
        actions.pack(fill=BOTH)

        ttk.Button(actions, text="Реєстри", command=self.open_registers).pack(side=LEFT, padx=8, pady=8, ipadx=24, ipady=8)
        self.transactions_button = ttk.Button(
            actions,
            text="Оновити історію транзакцій",
            command=self.run_transactions_threaded,
        )
        self.transactions_button.pack(side=LEFT, padx=8, pady=8, ipadx=24, ipady=8)

        balances_box = ttk.LabelFrame(main, text="Залишки на рахунках", padding=12)
        balances_box.pack(fill=BOTH, expand=True, pady=(14, 0))

        self.balances_table = ttk.Treeview(
            balances_box,
            columns=("bank", "currency", "balance", "date"),
            show="headings",
            height=4,
        )
        self.balances_table.heading("bank", text="Банк")
        self.balances_table.heading("currency", text="Валюта")
        self.balances_table.heading("balance", text="Залишок")
        self.balances_table.heading("date", text="Дата")
        self.balances_table.column("bank", width=150, anchor="w")
        self.balances_table.column("currency", width=80, anchor="center")
        self.balances_table.column("balance", width=160, anchor="e")
        self.balances_table.column("date", width=130, anchor="center")
        self.balances_table.pack(fill=BOTH, expand=True)

        self.balance_status = ttk.Label(balances_box, text="")
        self.balance_status.pack(fill=BOTH, pady=(8, 0))

    def open_registers(self) -> None:
        if self.registers_window is not None and self.registers_window.winfo_exists():
            self.registers_window.lift()
            self.registers_window.focus_force()
            return

        self.registers_window = Toplevel(self.root)
        RejestrApp(self.registers_window)

    def run_transactions_threaded(self) -> None:
        self.transactions_button.configure(state="disabled", text="Оновлюю...")
        thread = threading.Thread(target=self.run_transactions, daemon=True)
        thread.start()

    def run_transactions(self) -> None:
        output = io.StringIO()
        try:
            with contextlib.redirect_stdout(output), contextlib.redirect_stderr(output):
                transaction_processing.main()
            self.root.after(0, self._transactions_done)
        except Exception as exc:
            self.root.after(0, self._transactions_failed, str(exc))

    def _transactions_done(self) -> None:
        self.transactions_button.configure(state="normal", text="Оновити історію транзакцій")
        self.refresh_balances()
        messagebox.showinfo("Готово", "Історію транзакцій оновлено.")

    def _transactions_failed(self, error: str) -> None:
        self.transactions_button.configure(state="normal", text="Оновити історію транзакцій")
        messagebox.showerror("Помилка", error)

    def refresh_balances(self) -> None:
        for item in self.balances_table.get_children():
            self.balances_table.delete(item)

        try:
            balances = load_account_balances()
        except Exception as exc:
            self.balance_status.configure(text=f"Не вдалося прочитати залишки: {exc}")
            return

        for row in balances:
            self.balances_table.insert(
                "",
                "end",
                values=(
                    row["bank"],
                    row["currency"],
                    format_money(row["balance"], row["currency"]),
                    format_date(row["date"]),
                ),
            )

        self.balance_status.configure(text=f"Оновлено: {datetime.now().strftime('%H:%M:%S')}")


def load_account_balances() -> list[dict[str, Any]]:
    base_dir = transaction_processing.find_existing_base_dir()
    master_file = transaction_processing.find_master_file(base_dir)
    workbook = openpyxl.load_workbook(master_file, data_only=True, read_only=True)

    balances = []
    try:
        for bank, currency, sheet_name in BALANCE_SHEETS:
            if sheet_name not in workbook.sheetnames:
                balances.append({"bank": bank, "currency": currency, "balance": None, "date": None})
                continue

            ws = workbook[sheet_name]
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            header_index = {header: idx for idx, header in enumerate(headers) if header}
            date_idx = header_index.get("Дата транзакції")
            balance_idx = header_index.get("Сума після транзакції")

            balance = None
            tx_date = None
            if date_idx is not None and balance_idx is not None:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or all(value is None or str(value).strip() == "" for value in row):
                        continue
                    tx_date = row[date_idx]
                    balance = row[balance_idx]
                    break

            balances.append({"bank": bank, "currency": currency, "balance": balance, "date": tx_date})
    finally:
        workbook.close()

    return balances


def format_money(value: Any, currency: str) -> str:
    if value is None or value == "":
        return "—"
    try:
        number = float(value)
    except (TypeError, ValueError):
        return str(value)
    text = f"{number:,.2f}".replace(",", " ").replace(".", ",")
    return f"{text} {currency}"


def format_date(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if value is None or value == "":
        return "—"
    return str(value)
